"""Genera el Excel de revisión (legible) + la hoja de asientos, y el TXT plano para ContaI."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY, PETRO, GOLD, MINT = '172A37', '14465B', 'FFC701', 'D8E8E2'
WARN, NEW, NOTA = 'FFF3C4', 'FFE0B2', 'F0E6F5'


def generar_excel(asientos, resumen, nom_cuenta, ruta_salida, titulo='Revisión de compras'):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Revisión'
    thin = Side(style='thin', color='D9DEDE')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:I1')
    ws['A1'] = titulo
    ws['A1'].font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor=NAVY)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[1].height = 28
    ws.merge_cells('A2:I2')
    ws['A2'] = (f"{resumen['facturas']} facturas · Total ${resumen['total']:,.0f} · "
                f"{resumen['ok']} automáticas · {resumen['divididas']} divididas por XML · "
                f"{resumen['nuevos']} nuevos · {resumen['revisar']} a revisar · "
                f"Descuadres: {resumen['descuadres']}")
    ws['A2'].font = Font(name='Arial', size=9, color=PETRO, italic=True)
    ws['A2'].alignment = Alignment(horizontal='left', indent=1)

    cols = ['Fecha', 'Tipo', 'Factura', 'NIT', 'Proveedor', 'Total', 'IVA', 'Cuentas asignadas', 'Estado']
    hr = 4
    for j, c in enumerate(cols, 1):
        cell = ws.cell(hr, j, c)
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=PETRO)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = bd
    ws.row_dimensions[hr].height = 22

    for i, a in enumerate(asientos):
        row = hr + 1 + i
        if 'NUEVO' in a['estado']:
            fill = NEW
        elif 'Revisar' in a['estado'] or 'sin XML' in a['estado']:
            fill = WARN
        elif a['tipo'] == 'NC':
            fill = NOTA
        elif a['dividida']:
            fill = 'E8F0EC'
        else:
            fill = 'FFFFFF' if i % 2 else 'F4F8F7'
        ctas = ', '.join(f"{l['cuenta']}" for l in a['lineas'] if l['cuenta'])
        vals = [a['fecha'], a['tipo'], a['factura'], a['nit'], a['proveedor'],
                a['total'], a['iva'], ctas, a['estado']]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row, j, v)
            cell.font = Font(name='Arial', size=9, color=NAVY)
            cell.fill = PatternFill('solid', fgColor=fill)
            cell.border = bd
            if j in (6, 7):
                cell.number_format = '#,##0;(#,##0);-'
                cell.alignment = Alignment(horizontal='right')
            elif j in (2, 4):
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left', indent=1)
            if j == 9:
                col = '1F7A52' if v == 'OK' or 'Dividida' in str(v) else '9A6B00'
                cell.font = Font(name='Arial', size=9, bold=True, color=col)
    widths = [11, 6, 13, 12, 34, 13, 11, 24, 22]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A5'

    # Hoja de asientos (ContaExport)
    ws2 = wb.create_sheet('Asientos')
    h2 = ['Fecha', 'Factura', 'NIT', 'Proveedor', 'Cuenta', 'Nombre cuenta', 'Débito', 'Crédito']
    for j, c in enumerate(h2, 1):
        cell = ws2.cell(1, j, c)
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=PETRO)
        cell.border = bd
    rr = 2
    for a in asientos:
        for l in a['lineas']:
            ws2.cell(rr, 1, a['fecha'])
            ws2.cell(rr, 2, a['factura'])
            ws2.cell(rr, 3, a['nit'])
            ws2.cell(rr, 4, a['proveedor'])
            ws2.cell(rr, 5, l['cuenta'])
            ws2.cell(rr, 6, nom_cuenta.get(str(l['cuenta']), ''))
            ws2.cell(rr, 7, l['debito']).number_format = '#,##0.00'
            rr += 1
        ws2.cell(rr, 4, a['proveedor'])
        ws2.cell(rr, 5, a['prov_cuenta'])
        ws2.cell(rr, 6, nom_cuenta.get(str(a['prov_cuenta']), ''))
        ws2.cell(rr, 8, a['credito']).number_format = '#,##0.00'
        rr += 1
    for j, w in enumerate([11, 13, 12, 34, 11, 30, 14, 14], 1):
        ws2.column_dimensions[get_column_letter(j)].width = w
    ws2.freeze_panes = 'A2'

    wb.save(ruta_salida)
    return ruta_salida


def generar_txt(asientos, ruta_salida, nro_registro='00003'):
    """TXT plano de ContaI: 1=débito, 2=crédito. Documento agrupa el asiento.
    (Formato a confirmar contra una importación real de ContaI.)"""
    anchos = [20, 5, 10, 9, 9, 11, 28, 1, 21]
    lineas = []
    consec = 0
    for a in asientos:
        consec += 1
        doc = f"{consec:09d}"
        fecha = a['fecha'].replace('/', '/')  # DD/MM/YYYY -> se ajusta abajo
        d, m, y = a['fecha'].split('/')
        fecha = f"{m}/{d}/{y}"  # ContaI ingresos/egresos: MM/DD/YYYY
        for l in a['lineas']:
            if not l['cuenta']:
                continue
            campos = [str(l['cuenta']), nro_registro, fecha, doc, a['factura'],
                      a['nit'], a['proveedor'].upper(), '1', f"{abs(l['debito']):.2f}"]
            lineas.append('\t'.join(campos))
        campos = [str(a['prov_cuenta']), nro_registro, fecha, doc, a['factura'],
                  a['nit'], a['proveedor'].upper(), '2', f"{abs(a['credito']):.2f}"]
        lineas.append('\t'.join(campos))
    with open(ruta_salida, 'w', encoding='latin-1') as f:
        f.write('\n'.join(lineas))
    return ruta_salida
