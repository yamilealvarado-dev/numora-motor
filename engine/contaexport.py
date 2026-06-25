"""Exportación en formato ContaExport para ContaI.
TXT de ancho fijo (13 campos separados por TAB, cada campo pre-formateado con padding).
Replica exactamente el módulo Exportar documentado.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY, PETRO, MINT = '172A37', '14465B', 'D8E8E2'


# ---------- helpers de padding ----------
def pad_right(s, n):
    s = str(s or '')
    return s[:n].ljust(n, ' ')


def pad_left_ceros(s, n):
    s = str(s or '')
    return s[:n].rjust(n, '0')


# ---------- fecha MM/DD/YYYY ----------
def fecha_txt(fecha):
    """Recibe DD/MM/YYYY (formato interno) y devuelve MM/DD/YYYY (ancho 10)."""
    s = str(fecha or '').strip()
    if '/' in s:
        p = s.split('/')
        if len(p) == 3:
            a, b, c = p
            # interno es DD/MM/YYYY -> salida MM/DD/YYYY
            if len(a) <= 2 and len(c) == 4:
                return f"{b.zfill(2)}/{a.zfill(2)}/{c}"
            return s
    return s


# ---------- NIT: quitar no numéricos, CONSERVA dígito de verificación ----------
def limpiar_nit_txt(nit):
    import re
    return re.sub(r'\D', '', str(nit or ''))


# ---------- ¿la cuenta lleva NIT? ----------
def cuenta_lleva_nit(codigo, puc_tipos=None, override=None):
    """Regla Hi Día: TODAS las cuentas llevan NIT, EXCEPTO las que empiezan
    por 11 (disponible: caja y bancos). El override manual tiene prioridad."""
    codigo = str(codigo or '').replace('-', '').strip()
    if override is not None:
        return override
    if codigo.startswith('11'):
        return False
    return True


def _campo_nit(codigo, nit, puc_tipos, overrides):
    ov = (overrides or {}).get(codigo)
    if cuenta_lleva_nit(codigo, puc_tipos, ov):
        return pad_right(limpiar_nit_txt(nit), 11)
    return ' ' * 11


def _linea_txt(codigo, comprobante, fecha, documento, docref, nit, detalle,
               tipo, valor, valor_base='', ccosto=''):
    """Arma una línea de 13 campos con ancho fijo, unidos por TAB."""
    campos = [
        pad_right(str(codigo).replace('-', ''), 20),       # 1 código (20)
        pad_left_ceros(comprobante, 5),                    # 2 comprobante (5)
        fecha_txt(fecha)[:10].ljust(10),                   # 3 fecha (10)
        pad_right(documento, 9),                           # 4 documento (9)
        pad_right(docref, 9),                              # 5 doc referencia (9)
        nit,                                               # 6 NIT (ya viene a 11)
        pad_right(detalle, 28),                            # 7 detalle (28)
        str(tipo),                                         # 8 tipo 1/2 (1)
        pad_right(f"{abs(valor):.2f}", 21),                # 9 valor (21)
        pad_right(valor_base, 21),                         # 10 valor base (21)
        pad_right(ccosto, 20),                             # 11 centro costos (20)
        '   ',                                             # 12 transacción (3)
        '    ',                                            # 13 plazo (4)
    ]
    return '\t'.join(campos)


def _doc9(documento):
    """Últimos 9 del documento, sin guiones; si más corto, espacios a la derecha."""
    raw = str(documento or '').replace('-', '').strip()
    return raw[-9:] if len(raw) >= 9 else pad_right(raw, 9)


def generar_txt(asientos, ruta_salida, puc_tipos=None, nit_overrides=None,
                comprobante_por_tipo=None):
    """Genera el TXT ContaExport (ancho fijo). Cada asiento agrupa por comprobante+documento.
    comprobante_por_tipo: dict opcional {tipo_doc: comprobante} si cada factura trae su 'tipo_doc'."""
    lineas = []
    consec = 0
    for a in asientos:
        consec += 1
        # comprobante: el de la factura, o el del tipo, o 00003 por defecto
        comp = a.get('comprobante') or '00003'
        if comprobante_por_tipo and a.get('tipo_doc'):
            comp = comprobante_por_tipo.get(a['tipo_doc'], comp)
        doc = _doc9(a.get('factura') or f"{consec:09d}")
        docref = doc
        fecha = a.get('fecha', '')
        detalle = str(a.get('proveedor', '')).upper()

        # débitos
        for l in a['lineas']:
            cod = str(l.get('cuenta') or '')
            if not cod:
                continue
            nit_campo = _campo_nit(cod, a.get('nit'), puc_tipos, nit_overrides)
            vbase = l.get('valor_base', '') or ''
            cc = l.get('ccosto', '') or ''
            lineas.append(_linea_txt(cod, comp, fecha, doc, docref, nit_campo,
                                     detalle, '1', l['debito'], vbase, cc))
        # crédito al proveedor
        cod = str(a.get('prov_cuenta') or '')
        if cod:
            nit_campo = _campo_nit(cod, a.get('nit'), puc_tipos, nit_overrides)
            lineas.append(_linea_txt(cod, comp, fecha, doc, docref, nit_campo,
                                     detalle, '2', a['credito']))
    with open(ruta_salida, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lineas))
    return ruta_salida


def generar_excel_contaexport(asientos, nom_cuenta, ruta_salida, hoja='Datos',
                              puc_tipos=None, nit_overrides=None):
    """Excel intermedio ContaExport: mismas 13 columnas, cabecera en fila 1, datos desde fila 2."""
    wb = Workbook()
    ws = wb.active
    ws.title = hoja
    thin = Side(style='thin', color='D9DEDE')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    cols = ['Codigo Cuenta', 'Comprobante', 'Fecha', 'Documento', 'Documento Referencia',
            'Nit', 'Detalle', 'Tipo', 'Valor', 'Valor Base', 'Centro de costos',
            'Transaccion', 'Plazo']
    for j, c in enumerate(cols, 1):
        cell = ws.cell(1, j, c)
        cell.font = Font(name='Arial', bold=True, size=9, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=PETRO)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = bd

    r = 2
    for consec, a in enumerate(asientos, 1):
        comp = pad_left_ceros(a.get('comprobante') or '00003', 5)
        doc = _doc9(a.get('factura') or f"{consec:09d}")
        fecha = fecha_txt(a.get('fecha', ''))
        detalle = str(a.get('proveedor', '')).upper()[:28]
        for l in a['lineas']:
            cod = str(l.get('cuenta') or '')
            if not cod:
                continue
            nit_v = limpiar_nit_txt(a.get('nit')) if cuenta_lleva_nit(cod, puc_tipos, (nit_overrides or {}).get(cod)) else ''
            _fila(ws, r, bd, cod, comp, fecha, doc, doc, nit_v, detalle, '1',
                  l['debito'], l.get('valor_base', ''), l.get('ccosto', ''))
            r += 1
        cod = str(a.get('prov_cuenta') or '')
        if cod:
            nit_v = limpiar_nit_txt(a.get('nit')) if cuenta_lleva_nit(cod, puc_tipos, (nit_overrides or {}).get(cod)) else ''
            _fila(ws, r, bd, cod, comp, fecha, doc, doc, nit_v, detalle, '2',
                  a['credito'], '', '')
            r += 1

    widths = [22, 8, 12, 12, 12, 14, 30, 5, 22, 22, 22, 6, 6]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = 'A2'
    wb.save(ruta_salida)
    return ruta_salida


def _fila(ws, r, bd, codigo, comp, fecha, doc, docref, nit, detalle, tipo, valor, vbase, cc):
    vals = [codigo, comp, fecha, doc, docref, nit, detalle, tipo,
            f"{abs(valor):.2f}", vbase or '', cc or '', '', '']
    for j, v in enumerate(vals, 1):
        cell = ws.cell(r, j, v)
        cell.font = Font(name='Arial', size=9)
        cell.border = bd
        if j in (6, 4, 5):
            cell.number_format = '@'
        if j == 9:
            cell.alignment = Alignment(horizontal='right')
        else:
            cell.alignment = Alignment(horizontal='left', indent=1)
