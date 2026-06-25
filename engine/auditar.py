"""Auditoría: cruza las compras YA contabilizadas (auxiliar) contra el reporte DIAN.
Detecta: facturas en DIAN que faltan por contabilizar, contabilizadas sin soporte DIAN,
y diferencias de valor."""
import pandas as pd
from .aprender import limpiar_nit, _es_proveedor


def _num(x):
    return pd.to_numeric(x, errors='coerce').fillna(0)


def _folio_norm(x):
    """ContaI guarda los últimos 9 caracteres y rellena con ceros adelante (ej. MTC95852 -> 0MTC95852).
    Normaliza igual en ambos lados: últimos 9 y sin ceros de relleno."""
    import re
    s = re.sub(r'\s+', '', str(x or '').strip().upper())
    s = s[-9:] if len(s) > 9 else s
    return s.lstrip('0')


def meses_disponibles(ruta_dian):
    """Devuelve la lista de meses (YYYY-MM) con compras recibidas en el reporte DIAN."""
    df = pd.read_excel(ruta_dian, header=0, dtype=str)
    rec = df[df['Grupo'].astype(str).str.contains('ecib', na=False)].copy()
    fe = pd.to_datetime(rec['Fecha Emisión'], format='%d-%m-%Y', errors='coerce')
    return sorted(fe.dropna().dt.strftime('%Y-%m').unique().tolist())


def _filtrar_aux_mes(comp, mes):
    """Filtra el auxiliar al mes (YYYY-MM) por la columna Período o por Fecha."""
    if not mes:
        return comp
    yyyymm = mes.replace('-', '')
    if 'Período' in comp.columns and comp['Período'].notna().any():
        return comp[comp['Período'].astype(str).str.replace('-', '', regex=False) == yyyymm]
    fe = None
    for fmt in ('%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d'):
        fe = pd.to_datetime(comp['Fecha'], format=fmt, errors='coerce')
        if fe.notna().any():
            break
    return comp[fe.dt.strftime('%Y%m') == yyyymm]


def auditar(ruta_auxiliar, ruta_dian, comprobante='00003', mes=None):
    """Cruza compras contabilizadas vs DIAN. Si 'mes' (YYYY-MM), filtra ambos a ese mes."""
    # 1) Compras contabilizadas en el auxiliar
    aux = pd.read_excel(ruta_auxiliar, sheet_name='Datos', header=2, dtype=str)
    for c in ['Débitos', 'Créditos']:
        aux[c] = _num(aux[c])
    comp = aux[aux['Comprobante'] == comprobante].copy()
    comp = _filtrar_aux_mes(comp, mes)
    comp['nit'] = comp['NIT'].apply(limpiar_nit)

    contab = {}  # (nit, folio) -> {total, doc}
    for doc, a in comp.groupby('Documento'):
        pl = a[a['Cuenta'].apply(_es_proveedor) & (a['Créditos'] > 0)]
        if len(pl) == 0:
            continue
        nit = pl.iloc[0]['nit']
        folio = _folio_norm(pl.iloc[0].get('Documento Ref.') or '')
        total_deb = round(a['Débitos'].sum(), 2)
        contab[(nit, folio)] = {'total': total_deb, 'doc': doc,
                                'nombre': pl.iloc[0].get('Nombre NIT', '')}

    # 2) Recibidas según la DIAN
    df = pd.read_excel(ruta_dian, header=0, dtype=str)
    rec = df[df['Grupo'].astype(str).str.contains('ecib', na=False)].copy()
    rec['Total'] = _num(rec['Total'])
    rec['nit'] = rec['NIT Emisor'].apply(limpiar_nit)
    rec['folio_n'] = (rec['Prefijo'].fillna('') + rec['Folio'].fillna('')).apply(_folio_norm)
    if mes:
        fe = pd.to_datetime(rec['Fecha Emisión'], format='%d-%m-%Y', errors='coerce')
        rec = rec[fe.dt.strftime('%Y-%m') == mes]

    dian = {}
    for _, r in rec.iterrows():
        dian[(r['nit'], r['folio_n'])] = {
            'total': round(r['Total'], 2),
            'folio': f"{r['Prefijo'] or ''}{r['Folio'] or ''}",
            'nombre': str(r['Nombre Emisor'])[:40],
            'fecha': r['Fecha Emisión']}

    faltan, sin_dian, diferencias, ok = [], [], [], 0
    for k, d in dian.items():
        if k in contab:
            if abs(d['total'] - contab[k]['total']) > 100:  # tolerancia $100
                diferencias.append({'nit': k[0], 'folio': d['folio'], 'nombre': d['nombre'],
                                    'total_dian': d['total'], 'total_contab': contab[k]['total'],
                                    'diferencia': round(d['total'] - contab[k]['total'], 2)})
            else:
                ok += 1
        else:
            faltan.append({'nit': k[0], 'folio': d['folio'], 'nombre': d['nombre'],
                           'total': d['total'], 'fecha': d['fecha']})
    for k, c in contab.items():
        if k not in dian:
            sin_dian.append({'nit': k[0], 'folio': k[1], 'nombre': c['nombre'], 'total': c['total']})

    resumen = {
        'dian_total': len(dian), 'contabilizadas': len(contab), 'cruzan_ok': ok,
        'faltan_contabilizar': len(faltan), 'sin_dian': len(sin_dian),
        'diferencias_valor': len(diferencias),
    }
    return {'resumen': resumen, 'faltan': faltan, 'sin_dian': sin_dian, 'diferencias': diferencias}


def generar_excel_auditoria(audit, ruta_salida):
    """Excel de auditoría con 3 secciones: faltan, sin DIAN, diferencias."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    NAVY, PETRO, MINT = '172A37', '14465B', 'D8E8E2'
    WARN, BAD = 'FFF3C4', 'FCE4E2'
    thin = Side(style='thin', color='D9DEDE')
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Auditoría'
    r = audit['resumen']
    ws.merge_cells('A1:E1')
    ws['A1'] = 'Auditoría compras: contabilizado vs DIAN'
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor=NAVY)
    ws['A1'].alignment = Alignment(indent=1, vertical='center')
    ws.row_dimensions[1].height = 26
    ws.merge_cells('A2:E2')
    ws['A2'] = (f"DIAN: {r['dian_total']} · Contabilizadas: {r['contabilizadas']} · "
                f"Cruzan OK: {r['cruzan_ok']} · Faltan: {r['faltan_contabilizar']} · "
                f"Sin DIAN: {r['sin_dian']} · Diferencias: {r['diferencias_valor']}")
    ws['A2'].font = Font(size=9, italic=True, color=PETRO)
    ws['A2'].alignment = Alignment(indent=1)
    row = 4

    def seccion(titulo, datos, cols, fill):
        nonlocal row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(cols))
        ws.cell(row, 1, titulo).font = Font(bold=True, size=10, color='FFFFFF')
        ws.cell(row, 1).fill = PatternFill('solid', fgColor=PETRO)
        ws.cell(row, 1).alignment = Alignment(indent=1)
        row += 1
        for j, c in enumerate(cols, 1):
            cell = ws.cell(row, j, c)
            cell.font = Font(bold=True, size=9)
            cell.fill = PatternFill('solid', fgColor=MINT)
            cell.border = bd
        row += 1
        for d in datos:
            for j, key in enumerate(cols, 1):
                v = d.get(key.lower().replace(' ', '_'), '')
                cell = ws.cell(row, j, v)
                cell.font = Font(size=9)
                cell.fill = PatternFill('solid', fgColor=fill)
                cell.border = bd
                if isinstance(v, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
            row += 1
        row += 1

    seccion('FALTAN POR CONTABILIZAR (están en DIAN, no en el auxiliar)',
            audit['faltan'], ['nombre', 'folio', 'nit', 'total', 'fecha'], WARN)
    seccion('CONTABILIZADAS SIN DIAN (revisar: doc soporte, físicas o error)',
            audit['sin_dian'], ['nombre', 'folio', 'nit', 'total'], BAD)
    seccion('DIFERENCIAS DE VALOR (cruzan pero con monto distinto)',
            audit['diferencias'], ['nombre', 'folio', 'total_dian', 'total_contab', 'diferencia'], BAD)
    for j, w in enumerate([34, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    wb.save(ruta_salida)
    return ruta_salida
